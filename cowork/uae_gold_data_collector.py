"""
UAE Gold Trade & Dubai Premium/Discount - Historical Data Collector
===================================================================
Collects ~10 years of historical data (2016-present) from multiple sources
and outputs a multi-sheet Excel workbook with daily, monthly, and annual series.

Data sources (default mode):
  - Yahoo Finance: Gold, silver, FX rates, DXY, GLD ETF, US 10Y, VIX, crude oil
  - UN Comtrade API: UAE gold imports/exports by partner country (HS 7108)
  - CSV files in data/ folder: India duty, UAE CB reserves, SGE premium, Dubai
    premium, ETF holdings, India imports, Swiss/Turkey/Africa flows, CB purchases

Data sources (--bloomberg mode):
  - Bloomberg Terminal (blpapi): All daily market data, LBMA fixes, DGCX gold,
    SGE premium, Dubai premium, India gold imports, global ETF holdings, and more.
  - Falls back to Yahoo Finance / research estimates for any series that fail.
  - UN Comtrade is still used for partner-level trade data.

Proxy support:
  All HTTP requests (yfinance, Comtrade, etc.) can be routed through a proxy
  using --proxy http://host:port or --proxy socks5://host:port

Requirements:
  Default:    pip install yfinance pandas openpyxl requests
  Bloomberg:  pip install blpapi pandas openpyxl requests
  Proxy:      pip install requests[socks]   (only for SOCKS proxies)

Usage:
  python uae_gold_data_collector.py
  python uae_gold_data_collector.py --bloomberg
  python uae_gold_data_collector.py --proxy http://127.0.0.1:8080
  python uae_gold_data_collector.py --proxy socks5://127.0.0.1:1080
  python uae_gold_data_collector.py --data-dir /path/to/data
  python uae_gold_data_collector.py --start 2018-01-01 --end 2025-12-31
"""

import argparse
import os
import time
from datetime import datetime

import pandas as pd
import numpy as np
import requests
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Optional imports — loaded on demand
yf = None
blpapi = None

# Global proxy dict — set from CLI via configure_proxy()
_PROXY_DICT = {}


# ============================================================
# CONFIG
# ============================================================
DEFAULT_START = "2016-01-01"
DEFAULT_END = datetime.now().strftime("%Y-%m-%d")
DEFAULT_OUTPUT = "UAE_Gold_Trade_Historical_Data.xlsx"
DEFAULT_DATA_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'data')

# Expected data files (filename -> description)
DATA_FILES = {
    'india_gold_duty.csv':          'India gold import duty timeline',
    'uae_cb_gold_reserves.csv':     'UAE Central Bank gold reserves',
    'sge_premium_estimate.csv':     'Shanghai Gold Exchange premium estimates',
    'dubai_premium_estimate.csv':   'Dubai gold premium/discount estimates',
    'gold_etf_holdings_estimate.csv': 'Global gold ETF holdings estimates',
    'india_gold_imports_estimate.csv': 'India gold import estimates',
    'swiss_gold_to_uae.csv':        'Swiss gold exports to UAE',
    'turkey_uae_gold.csv':          'Turkey-UAE gold trade',
    'africa_gold_to_uae.csv':       'African gold exports to UAE',
    'global_cb_gold_purchases.csv': 'Global central bank gold purchases',
}

M49_COUNTRY_CODES = {
    0: 'World', 4: 'Afghanistan', 12: 'Algeria', 24: 'Angola', 31: 'Azerbaijan',
    36: 'Australia', 40: 'Austria', 48: 'Bahrain', 50: 'Bangladesh', 51: 'Armenia',
    56: 'Belgium', 68: 'Bolivia', 76: 'Brazil', 100: 'Bulgaria', 108: 'Burundi',
    116: 'Cambodia', 120: 'Cameroon', 124: 'Canada', 140: 'Central African Rep.',
    144: 'Sri Lanka', 148: 'Chad', 152: 'Chile', 156: 'China', 170: 'Colombia',
    174: 'Comoros', 178: 'Congo', 180: 'DR Congo', 196: 'Cyprus', 204: 'Benin',
    208: 'Denmark', 218: 'Ecuador', 222: 'El Salvador', 231: 'Ethiopia',
    233: 'Estonia', 246: 'Finland', 250: 'France', 268: 'Georgia', 276: 'Germany',
    288: 'Ghana', 300: 'Greece', 324: 'Guinea', 332: 'Haiti', 344: 'Hong Kong',
    348: 'Hungary', 356: 'India', 360: 'Indonesia', 364: 'Iran', 368: 'Iraq',
    372: 'Ireland', 376: 'Israel', 380: 'Italy', 384: "Cote d'Ivoire",
    392: 'Japan', 398: 'Kazakhstan', 400: 'Jordan', 404: 'Kenya',
    410: 'South Korea', 414: 'Kuwait', 418: 'Laos', 422: 'Lebanon',
    434: 'Libya', 442: 'Luxembourg', 450: 'Madagascar', 454: 'Malawi',
    458: 'Malaysia', 466: 'Mali', 478: 'Mauritania', 480: 'Mauritius',
    484: 'Mexico', 496: 'Mongolia', 504: 'Morocco', 508: 'Mozambique',
    512: 'Oman', 516: 'Namibia', 524: 'Nepal', 528: 'Netherlands',
    554: 'New Zealand', 562: 'Niger', 566: 'Nigeria', 578: 'Norway',
    586: 'Pakistan', 604: 'Peru', 608: 'Philippines', 616: 'Poland',
    620: 'Portugal', 634: 'Qatar', 642: 'Romania', 643: 'Russia',
    646: 'Rwanda', 682: 'Saudi Arabia', 686: 'Senegal', 694: 'Sierra Leone',
    699: 'India', 702: 'Singapore', 710: 'South Africa', 716: 'Zimbabwe',
    724: 'Spain', 729: 'Sudan', 740: 'Suriname', 756: 'Switzerland',
    757: 'Switzerland', 760: 'Syria', 764: 'Thailand', 768: 'Togo',
    780: 'Trinidad and Tobago', 784: 'UAE', 788: 'Tunisia', 792: 'Turkey',
    800: 'Uganda', 804: 'Ukraine', 818: 'Egypt', 826: 'United Kingdom',
    834: 'Tanzania', 840: 'United States', 854: 'Burkina Faso', 858: 'Uruguay',
    860: 'Uzbekistan', 862: 'Venezuela', 887: 'Yemen',
}

# ============================================================
# BLOOMBERG TICKER MAP
# ============================================================
BBG_DAILY_TICKERS = {
    'COMEX_Gold_Close_USD':   ('GC1 Comdty', 'PX_LAST'),
    'LBMA_Gold_AM_USD':       ('GOLDLNAM Index', 'PX_LAST'),
    'LBMA_Gold_PM_USD':       ('GOLDLNPM Index', 'PX_LAST'),
    'DGCX_Gold_Close_USD':    ('OG1 DGCX Comdty', 'PX_LAST'),
    'Silver_Close_USD':       ('SI1 Comdty', 'PX_LAST'),
    'Gold_AED_oz':            ('XAUAED Curncy', 'PX_LAST'),
    'USD_INR':                ('USDINR Curncy', 'PX_LAST'),
    'USD_CNY':                ('USDCNY Curncy', 'PX_LAST'),
    'USD_TRY':                ('USDTRY Curncy', 'PX_LAST'),
    'USD_AED':                ('USDAED Curncy', 'PX_LAST'),
    'DXY_Index':              ('DXY Curncy', 'PX_LAST'),
    'GLD_Close':              ('GLD US Equity', 'PX_LAST'),
    'GLD_Volume':             ('GLD US Equity', 'PX_VOLUME'),
    'US_10Y_Yield':           ('USGG10YR Index', 'PX_LAST'),
    'VIX':                    ('VIX Index', 'PX_LAST'),
    'WTI_Crude_USD':          ('CL1 Comdty', 'PX_LAST'),
    'Brent_Crude_USD':        ('CO1 Comdty', 'PX_LAST'),
    'US_2Y_Yield':            ('USGG2YR Index', 'PX_LAST'),
    'US_Real_Rate_10Y':       ('USGGT10Y Index', 'PX_LAST'),
    'Fed_Funds_Rate':         ('FDTRMID Index', 'PX_LAST'),
}

BBG_MONTHLY_TICKERS = {
    'SGE_Premium_USD_oz':              ('GLDPSGP Index', 'PX_LAST'),
    'Dubai_Premium_USD_oz':            ('GLDPDXB Index', 'PX_LAST'),
    'Global_Gold_ETF_Holdings_Tonnes': ('TGOLDTOT Index', 'PX_LAST'),
    'India_Gold_Imports_USD_Bn':       ('INGDIMPM Index', 'PX_LAST'),
    'Global_Gold_Mine_Supply_Tonnes':  ('GLDMNSUP Index', 'PX_LAST'),
    'Gold_COMEX_Open_Interest':        ('GCOINTNR Index', 'PX_LAST'),
    'Gold_CFTC_Net_Long':             ('CFAUNETL Index', 'PX_LAST'),
}


# ============================================================
# PROXY & LAZY IMPORTS
# ============================================================
def configure_proxy(proxy_url):
    """Set the global proxy dict used by requests and yfinance."""
    global _PROXY_DICT
    if proxy_url:
        _PROXY_DICT = {'http': proxy_url, 'https': proxy_url}
        print(f"  Proxy configured: {proxy_url}")
    else:
        _PROXY_DICT = {}


def _get_session():
    """Return a requests.Session with proxy configured (if any)."""
    s = requests.Session()
    if _PROXY_DICT:
        s.proxies.update(_PROXY_DICT)
    return s


def _ensure_yfinance():
    """Lazy-import yfinance and configure its proxy."""
    global yf
    if yf is None:
        import yfinance as _yf
        yf = _yf


def _ensure_blpapi():
    """Lazy-import blpapi and return the module."""
    global blpapi
    if blpapi is None:
        import blpapi as _blp
        blpapi = _blp
    return blpapi


# ============================================================
# BLOOMBERG DATA FUNCTIONS
# ============================================================
def _bbg_session():
    """Open a blpapi Session connected to localhost:8194."""
    _ensure_blpapi()
    opts = blpapi.SessionOptions()
    opts.setServerHost('localhost')
    opts.setServerPort(8194)
    session = blpapi.Session(opts)
    if not session.start():
        raise RuntimeError("Failed to start Bloomberg session. Is the Terminal running?")
    if not session.openService('//blp/refdata'):
        session.stop()
        raise RuntimeError("Failed to open //blp/refdata service.")
    return session


def _bbg_bdh(session, security, field, start_date, end_date,
             periodicity='DAILY', overrides=None):
    """Run a Bloomberg BDH (HistoricalDataRequest) and return a DataFrame.

    Parameters
    ----------
    session : blpapi.Session
    security : str   e.g. "GC1 Comdty"
    field : str      e.g. "PX_LAST"
    start_date : str YYYYMMDD
    end_date : str   YYYYMMDD
    periodicity : str  DAILY | MONTHLY | YEARLY
    overrides : dict   optional {field: value} overrides

    Returns
    -------
    pd.DataFrame with DatetimeIndex ('Date') and one column named `field`.
    Empty DataFrame on failure.
    """
    refdata = session.getService('//blp/refdata')
    req = refdata.createRequest('HistoricalDataRequest')
    req.getElement('securities').appendValue(security)
    req.getElement('fields').appendValue(field)
    req.set('startDate', start_date)
    req.set('endDate', end_date)
    req.set('periodicitySelection', periodicity)
    req.set('nonTradingDayFillOption', 'ACTIVE_DAYS_ONLY')
    if overrides:
        ov_elem = req.getElement('overrides')
        for k, v in overrides.items():
            o = ov_elem.appendElement()
            o.setElement('fieldId', k)
            o.setElement('value', str(v))

    session.sendRequest(req)

    dates, values = [], []
    while True:
        ev = session.nextEvent(500)
        for msg in ev:
            if msg.hasElement('securityData'):
                sec_data = msg.getElement('securityData')
                if sec_data.hasElement('fieldData'):
                    fd_array = sec_data.getElement('fieldData')
                    for i in range(fd_array.numValues()):
                        pt = fd_array.getValueAsElement(i)
                        dates.append(pt.getElementAsDatetime('date'))
                        try:
                            values.append(pt.getElementAsFloat(field))
                        except Exception:
                            values.append(np.nan)
        if ev.eventType() == blpapi.Event.RESPONSE:
            break

    if not dates:
        return pd.DataFrame()

    idx = pd.to_datetime(dates)
    return pd.DataFrame({field: values}, index=idx).rename_axis('Date')


def collect_daily_market_data_bloomberg(start, end):
    """Fetch all daily market series from Bloomberg Terminal via blpapi.

    Falls back to Yahoo Finance per-series on failure.
    """
    print("\n" + "=" * 60)
    print("STEP 1: DAILY MARKET DATA (Bloomberg Terminal)")
    print("=" * 60)

    start_str = pd.Timestamp(start).strftime('%Y%m%d')
    end_str = pd.Timestamp(end).strftime('%Y%m%d')

    try:
        session = _bbg_session()
    except Exception as e:
        print(f"  Bloomberg connection failed: {e}")
        print("  Falling back to Yahoo Finance for all daily data.")
        return collect_daily_market_data(start, end), 'Yahoo Finance (Bloomberg unavailable)'

    frames = []
    bbg_sourced = []
    yf_fallback = []

    for col_name, (ticker, field) in BBG_DAILY_TICKERS.items():
        print(f"  {col_name} [{ticker}]...", end=" ")
        try:
            df = _bbg_bdh(session, ticker, field, start_str, end_str, 'DAILY')
            if len(df) > 0:
                df = df.rename(columns={field: col_name})
                frames.append(df)
                bbg_sourced.append(col_name)
                print(f"{len(df)} rows")
            else:
                raise ValueError("empty result")
        except Exception as e:
            print(f"FAILED ({e}) — will try YF fallback")
            yf_fallback.append(col_name)

    session.stop()

    # Merge Bloomberg frames
    if frames:
        daily = frames[0]
        for f in frames[1:]:
            daily = daily.join(f, how='outer')
    else:
        daily = pd.DataFrame()

    # Yahoo Finance fallback for any missing series
    if yf_fallback:
        _ensure_yfinance()
        YF_MAP = {
            'COMEX_Gold_Close_USD': ('GC=F', 'COMEX_Gold_Close_USD'),
            'Silver_Close_USD': ('SI=F', 'Silver_Close_USD'),
            'USD_INR': ('INR=X', 'USD_INR'),
            'USD_CNY': ('CNY=X', 'USD_CNY'),
            'USD_TRY': ('TRY=X', 'USD_TRY'),
            'DXY_Index': ('DX-Y.NYB', 'DXY_Index'),
            'GLD_Close': ('GLD', 'GLD_Close'),
            'GLD_Volume': ('GLD', 'GLD_Volume'),
            'US_10Y_Yield': ('^TNX', 'US_10Y_Yield'),
            'VIX': ('^VIX', 'VIX'),
            'WTI_Crude_USD': ('CL=F', 'WTI_Crude_USD'),
        }
        for col_name in yf_fallback:
            if col_name in YF_MAP:
                ticker, name = YF_MAP[col_name]
                print(f"  [YF fallback] {col_name} [{ticker}]...", end=" ")
                s = fetch_yf_series(ticker, name, start, end)
                if len(s) > 0:
                    if len(daily) > 0:
                        daily = daily.join(s, how='outer')
                    else:
                        daily = s

    daily.index = pd.to_datetime(daily.index)
    daily.index.name = 'Date'
    daily = daily.sort_index()
    daily = daily[daily.index >= start]

    # Computed columns
    if 'COMEX_Gold_Close_USD' in daily.columns and 'Silver_Close_USD' in daily.columns:
        daily['Gold_Silver_Ratio'] = daily['COMEX_Gold_Close_USD'] / daily['Silver_Close_USD']
    if 'COMEX_Gold_Close_USD' in daily.columns and 'Gold_AED_oz' not in daily.columns:
        daily['Gold_AED_oz'] = daily['COMEX_Gold_Close_USD'] * 3.6725
    if 'COMEX_Gold_Close_USD' in daily.columns and 'USD_INR' in daily.columns:
        daily['Gold_INR_per_10g'] = daily['COMEX_Gold_Close_USD'] * daily['USD_INR'] / 31.1035 * 10
    if 'US_10Y_Yield' in daily.columns and 'US_2Y_Yield' in daily.columns:
        daily['US_Yield_Curve_2s10s'] = daily['US_10Y_Yield'] - daily['US_2Y_Yield']
    if 'LBMA_Gold_PM_USD' in daily.columns and 'COMEX_Gold_Close_USD' in daily.columns:
        daily['LBMA_COMEX_Spread'] = daily['LBMA_Gold_PM_USD'] - daily['COMEX_Gold_Close_USD']

    print(f"\n  => Daily merged: {daily.shape[0]} rows x {daily.shape[1]} columns")
    print(f"  => Bloomberg sourced: {len(bbg_sourced)} series")
    print(f"  => YF fallback: {len([c for c in yf_fallback if c in (daily.columns if len(daily) > 0 else [])])} series")

    source_label = 'Bloomberg Terminal' if bbg_sourced else 'Yahoo Finance'
    return daily, source_label


def collect_monthly_bloomberg_data(start, end):
    """Fetch monthly series from Bloomberg (SGE premium, Dubai premium, ETF holdings, etc.)."""
    print("\n  [Bloomberg monthly series]")
    start_str = pd.Timestamp(start).strftime('%Y%m%d')
    end_str = pd.Timestamp(end).strftime('%Y%m%d')

    results = {}
    try:
        session = _bbg_session()
    except Exception as e:
        print(f"  Bloomberg session failed: {e}")
        return results

    for col_name, (ticker, field) in BBG_MONTHLY_TICKERS.items():
        print(f"    {col_name} [{ticker}]...", end=" ")
        try:
            df = _bbg_bdh(session, ticker, field, start_str, end_str, 'MONTHLY')
            if len(df) > 0:
                df = df.rename(columns={field: col_name})
                results[col_name] = df
                print(f"{len(df)} rows")
            else:
                print("empty")
        except Exception as e:
            print(f"FAILED ({e})")

    session.stop()
    return results


# ============================================================
# STEP 1: DAILY MARKET DATA (Yahoo Finance)
# ============================================================
def fetch_yf_series(ticker, col_name, start, end):
    """Download a single Yahoo Finance series and return a clean DataFrame."""
    _ensure_yfinance()
    try:
        # Pass proxy to yfinance via session
        session_obj = None
        if _PROXY_DICT:
            session_obj = _get_session()
        df = yf.download(ticker, start=start, end=end, progress=False,
                         session=session_obj)
        if 'Close' not in df.columns and len(df.columns) > 0:
            if hasattr(df.columns, 'get_level_values'):
                df.columns = df.columns.get_level_values(0)
        result = df[['Close']].rename(columns={'Close': col_name})
        result.index.name = 'Date'
        if hasattr(result.columns, 'droplevel'):
            try:
                result.columns = result.columns.droplevel(1)
            except (IndexError, KeyError):
                pass
        print(f"  {col_name}: {len(result)} rows")
        return result
    except Exception as e:
        print(f"  {col_name}: ERROR - {e}")
        return pd.DataFrame()


def fetch_yf_multi(ticker, col_map, start, end):
    """Download multiple columns from one Yahoo Finance ticker."""
    _ensure_yfinance()
    try:
        session_obj = None
        if _PROXY_DICT:
            session_obj = _get_session()
        df = yf.download(ticker, start=start, end=end, progress=False,
                         session=session_obj)
        if hasattr(df.columns, 'get_level_values'):
            try:
                df.columns = df.columns.get_level_values(0)
            except Exception:
                pass
        result = df[list(col_map.keys())].rename(columns=col_map)
        result.index.name = 'Date'
        if hasattr(result.columns, 'droplevel'):
            try:
                result.columns = result.columns.droplevel(1)
            except (IndexError, KeyError):
                pass
        print(f"  {list(col_map.values())}: {len(result)} rows")
        return result
    except Exception as e:
        print(f"  {list(col_map.values())}: ERROR - {e}")
        return pd.DataFrame()


def collect_daily_market_data(start, end):
    """Fetch all daily market series and merge into one DataFrame."""
    print("\n" + "=" * 60)
    print("STEP 1: DAILY MARKET DATA (Yahoo Finance)")
    print("=" * 60)

    series = [
        fetch_yf_series("GC=F", "COMEX_Gold_Close_USD", start, end),
        fetch_yf_series("SI=F", "Silver_Close_USD", start, end),
        fetch_yf_series("INR=X", "USD_INR", start, end),
        fetch_yf_series("CNY=X", "USD_CNY", start, end),
        fetch_yf_series("TRY=X", "USD_TRY", start, end),
        fetch_yf_series("DX-Y.NYB", "DXY_Index", start, end),
        fetch_yf_multi("GLD", {'Close': 'GLD_Close', 'Volume': 'GLD_Volume'}, start, end),
        fetch_yf_series("^TNX", "US_10Y_Yield", start, end),
        fetch_yf_series("^VIX", "VIX", start, end),
        fetch_yf_series("CL=F", "WTI_Crude_USD", start, end),
    ]

    series = [s for s in series if len(s) > 0]
    for s in series:
        s.index = pd.to_datetime(s.index)

    daily = series[0]
    for s in series[1:]:
        daily = daily.join(s, how='outer')

    daily.index.name = 'Date'
    daily = daily.sort_index()
    daily = daily[daily.index >= start]

    # Computed columns
    if 'COMEX_Gold_Close_USD' in daily.columns and 'Silver_Close_USD' in daily.columns:
        daily['Gold_Silver_Ratio'] = daily['COMEX_Gold_Close_USD'] / daily['Silver_Close_USD']
    if 'COMEX_Gold_Close_USD' in daily.columns:
        daily['Gold_AED_oz'] = daily['COMEX_Gold_Close_USD'] * 3.6725
    if 'COMEX_Gold_Close_USD' in daily.columns and 'USD_INR' in daily.columns:
        daily['Gold_INR_per_10g'] = daily['COMEX_Gold_Close_USD'] * daily['USD_INR'] / 31.1035 * 10

    print(f"\n  => Daily merged: {daily.shape[0]} rows x {daily.shape[1]} columns")
    return daily


# ============================================================
# STEP 2: UN COMTRADE - UAE GOLD TRADE (HS 7108)
# ============================================================
def _comtrade_fetch_periods(reporter_code='784', freq='A'):
    """Query the Comtrade data-availability endpoint and return available periods."""
    url = f"https://comtradeapi.un.org/public/v1/getDA/C/{freq}/HS"
    try:
        sess = _get_session()
        r = sess.get(url, params={'reporterCode': reporter_code}, timeout=30)
        if r.status_code == 200:
            data = r.json()
            return sorted(d.get('period') for d in data.get('data', []))
    except Exception:
        pass
    return []


def _comtrade_fetch_one(freq, period, flow, reporter='784', cmd='7108'):
    """Fetch a single period/flow combination from Comtrade. Returns list of dicts."""
    base_url = f"https://comtradeapi.un.org/public/v1/preview/C/{freq}/HS"
    params = {
        'reporterCode': reporter,
        'period': str(period),
        'cmdCode': cmd,
        'flowCode': flow,
    }
    try:
        sess = _get_session()
        r = sess.get(base_url, params=params, timeout=30)
        if r.status_code == 200:
            data = r.json()
            return data.get('data', [])
    except Exception:
        pass
    return []


def collect_comtrade_data(start_year, end_year):
    """Fetch UAE gold imports and exports from UN Comtrade public API.

    Pulls ANNUAL data for all available years, plus MONTHLY data for any
    months the UAE has reported (currently 2017-2019).
    """
    print("\n" + "=" * 60)
    print("STEP 2: UN COMTRADE - UAE GOLD TRADE (HS 7108)")
    print("=" * 60)

    # --- 2a. Annual data ---
    print("\n  [Annual data]")
    annual_imports_rec, annual_exports_rec = [], []

    for year in range(start_year, min(end_year + 1, datetime.now().year)):
        print(f"    {year}...", end=" ")
        for flow_code, records, val_col, wt_col, qty_col in [
            ('M', annual_imports_rec, 'Import_Value_USD', 'Import_NetWeight_Kg', 'Import_Qty'),
            ('X', annual_exports_rec, 'Export_Value_USD', 'Export_NetWeight_Kg', 'Export_Qty'),
        ]:
            rows = _comtrade_fetch_one('A', year, flow_code)
            for row in rows:
                code = row.get('partnerCode')
                records.append({
                    'Year': year,
                    'PartnerCode': code,
                    'Partner': M49_COUNTRY_CODES.get(code, f'Unknown_{code}'),
                    val_col: row.get('primaryValue'),
                    wt_col: row.get('netWgt'),
                    qty_col: row.get('qty'),
                })
            time.sleep(0.5)
        print("done")

    df_ann_imp = pd.DataFrame(annual_imports_rec) if annual_imports_rec else pd.DataFrame()
    df_ann_exp = pd.DataFrame(annual_exports_rec) if annual_exports_rec else pd.DataFrame()
    print(f"  => Annual imports: {len(df_ann_imp)} records")
    print(f"  => Annual exports: {len(df_ann_exp)} records")

    # --- 2b. Monthly data ---
    print("\n  [Monthly data - checking availability]")
    available_months = _comtrade_fetch_periods('784', 'M')
    available_months = [
        p for p in available_months
        if start_year <= int(str(p)[:4]) <= end_year
    ]
    print(f"  Available monthly periods for UAE (in range): {len(available_months)}")
    if available_months:
        yrs = sorted(set(str(p)[:4] for p in available_months))
        print(f"  Years with monthly data: {', '.join(yrs)}")

    monthly_imports_rec, monthly_exports_rec = [], []

    for i, period in enumerate(available_months):
        y = int(str(period)[:4])
        m = int(str(period)[4:6])
        print(f"    {period} ({i+1}/{len(available_months)})...", end=" ")

        for flow_code, records, val_col, wt_col, qty_col in [
            ('M', monthly_imports_rec, 'Import_Value_USD', 'Import_NetWeight_Kg', 'Import_Qty'),
            ('X', monthly_exports_rec, 'Export_Value_USD', 'Export_NetWeight_Kg', 'Export_Qty'),
        ]:
            rows = _comtrade_fetch_one('M', period, flow_code)
            for row in rows:
                code = row.get('partnerCode')
                records.append({
                    'Year': y,
                    'Month': m,
                    'Period': period,
                    'Date': pd.Timestamp(year=y, month=m, day=1),
                    'PartnerCode': code,
                    'Partner': M49_COUNTRY_CODES.get(code, f'Unknown_{code}'),
                    val_col: row.get('primaryValue'),
                    wt_col: row.get('netWgt'),
                    qty_col: row.get('qty'),
                })
            time.sleep(0.5)
        print("done")

    df_mon_imp = pd.DataFrame(monthly_imports_rec) if monthly_imports_rec else pd.DataFrame()
    df_mon_exp = pd.DataFrame(monthly_exports_rec) if monthly_exports_rec else pd.DataFrame()
    print(f"  => Monthly imports: {len(df_mon_imp)} records")
    print(f"  => Monthly exports: {len(df_mon_exp)} records")

    return df_ann_imp, df_ann_exp, df_mon_imp, df_mon_exp


# ============================================================
# STEP 3: FILE-BASED RESEARCH DATA (with Bloomberg overrides)
# ============================================================
def _load_csv(data_dir, filename, date_col='Date', parse_dates=True):
    """Load a CSV from the data directory. Returns DataFrame or empty DataFrame."""
    path = os.path.join(data_dir, filename)
    if not os.path.exists(path):
        print(f"  WARNING: {filename} not found in {data_dir}")
        return pd.DataFrame()
    try:
        if parse_dates and date_col:
            df = pd.read_csv(path, parse_dates=[date_col])
        else:
            df = pd.read_csv(path)
        return df
    except Exception as e:
        print(f"  WARNING: Failed to read {filename}: {e}")
        return pd.DataFrame()


def collect_research_data(data_dir, bbg_monthly=None):
    """Load research data from CSV files, with optional Bloomberg overrides.

    Parameters
    ----------
    data_dir : str
        Path to the data/ directory containing the CSV files.
    bbg_monthly : dict or None
        If provided (from Bloomberg), maps col_name -> pd.DataFrame.
        These replace the file-based data for those series.
    """
    if bbg_monthly is None:
        bbg_monthly = {}

    print("\n" + "=" * 60)
    print(f"STEP 3: RESEARCH DATA (from {data_dir})")
    print("=" * 60)

    # --- India Gold Import Duty ---
    india_duty = _load_csv(data_dir, 'india_gold_duty.csv')
    print(f"  India duty timeline: {len(india_duty)} entries")

    # --- UAE Central Bank Gold Reserves ---
    uae_reserves = _load_csv(data_dir, 'uae_cb_gold_reserves.csv')
    print(f"  UAE CB reserves: {len(uae_reserves)} entries")

    # --- SGE Premium (Bloomberg override or file) ---
    if 'SGE_Premium_USD_oz' in bbg_monthly:
        df_bbg = bbg_monthly['SGE_Premium_USD_oz'].copy()
        df_bbg.index = df_bbg.index.to_period('M').to_timestamp()
        sge_premium = df_bbg.reset_index().rename(columns={'index': 'Date'})
        sge_premium.columns = ['Date', 'SGE_Premium_USD_oz']
        sge_premium['Source'] = 'Bloomberg (GLDPSGP Index)'
        print(f"  SGE premium: {len(sge_premium)} months [Bloomberg]")
    else:
        sge_premium = _load_csv(data_dir, 'sge_premium_estimate.csv')
        print(f"  SGE premium: {len(sge_premium)} months [file]")

    # --- Dubai Premium (Bloomberg override or file) ---
    if 'Dubai_Premium_USD_oz' in bbg_monthly:
        df_bbg = bbg_monthly['Dubai_Premium_USD_oz'].copy()
        df_bbg.index = df_bbg.index.to_period('M').to_timestamp()
        dubai_premium = df_bbg.reset_index().rename(columns={'index': 'Date'})
        dubai_premium.columns = ['Date', 'Dubai_Premium_USD_oz']
        dubai_premium['Source'] = 'Bloomberg (GLDPDXB Index)'
        print(f"  Dubai premium: {len(dubai_premium)} months [Bloomberg]")
    else:
        dubai_premium = _load_csv(data_dir, 'dubai_premium_estimate.csv')
        print(f"  Dubai premium: {len(dubai_premium)} months [file]")

    # --- Gold ETF Holdings (Bloomberg override or file) ---
    if 'Global_Gold_ETF_Holdings_Tonnes' in bbg_monthly:
        df_bbg = bbg_monthly['Global_Gold_ETF_Holdings_Tonnes'].copy()
        df_bbg.index = df_bbg.index.to_period('M').to_timestamp()
        gold_etf = df_bbg.reset_index().rename(columns={'index': 'Date'})
        gold_etf.columns = ['Date', 'Global_Gold_ETF_Holdings_Tonnes']
        gold_etf['Source'] = 'Bloomberg (TGOLDTOT Index)'
        print(f"  Gold ETF holdings: {len(gold_etf)} months [Bloomberg]")
    else:
        gold_etf = _load_csv(data_dir, 'gold_etf_holdings_estimate.csv')
        print(f"  Gold ETF holdings: {len(gold_etf)} months [file]")

    # --- India Gold Imports (Bloomberg override or file) ---
    if 'India_Gold_Imports_USD_Bn' in bbg_monthly:
        df_bbg = bbg_monthly['India_Gold_Imports_USD_Bn'].copy()
        df_bbg.index = df_bbg.index.to_period('M').to_timestamp()
        india_gold_imports = df_bbg.reset_index().rename(columns={'index': 'Date'})
        india_gold_imports.columns = ['Date', 'India_Gold_Imports_USD_Bn']
        if india_gold_imports['India_Gold_Imports_USD_Bn'].median() > 100:
            india_gold_imports['India_Gold_Imports_USD_Bn'] /= 1000.0
        india_gold_imports['Source'] = 'Bloomberg (INGDIMPM Index)'
        print(f"  India gold imports: {len(india_gold_imports)} months [Bloomberg]")
    else:
        india_gold_imports = _load_csv(data_dir, 'india_gold_imports_estimate.csv')
        print(f"  India gold imports: {len(india_gold_imports)} months [file]")

    # --- Swiss Gold Exports to UAE ---
    swiss_to_uae = _load_csv(data_dir, 'swiss_gold_to_uae.csv', date_col=None, parse_dates=False)
    print(f"  Swiss->UAE gold: {len(swiss_to_uae)} years")

    # --- Turkey-UAE Gold Trade ---
    turkey_uae = _load_csv(data_dir, 'turkey_uae_gold.csv', date_col=None, parse_dates=False)
    print(f"  Turkey<->UAE gold: {len(turkey_uae)} years")

    # --- African Gold Exports to UAE ---
    africa_uae = _load_csv(data_dir, 'africa_gold_to_uae.csv', date_col=None, parse_dates=False)
    print(f"  Africa->UAE gold: {len(africa_uae)} records")

    # --- Global Central Bank Gold Purchases ---
    cb_purchases = _load_csv(data_dir, 'global_cb_gold_purchases.csv', date_col=None, parse_dates=False)
    print(f"  Global CB purchases: {len(cb_purchases)} years")

    # --- Extra Bloomberg-only monthly series ---
    bbg_extras = {}
    for col_name in ['Global_Gold_Mine_Supply_Tonnes', 'Gold_COMEX_Open_Interest', 'Gold_CFTC_Net_Long']:
        if col_name in bbg_monthly:
            df_bbg = bbg_monthly[col_name].copy()
            df_bbg.index = df_bbg.index.to_period('M').to_timestamp()
            df_bbg = df_bbg.reset_index().rename(columns={'index': 'Date'})
            df_bbg.columns = ['Date', col_name]
            bbg_extras[col_name] = df_bbg
            print(f"  {col_name}: {len(df_bbg)} months [Bloomberg]")

    return {
        'india_duty': india_duty,
        'uae_reserves': uae_reserves,
        'sge_premium': sge_premium,
        'dubai_premium': dubai_premium,
        'gold_etf': gold_etf,
        'india_gold_imports': india_gold_imports,
        'swiss_to_uae': swiss_to_uae,
        'turkey_uae': turkey_uae,
        'africa_uae': africa_uae,
        'cb_purchases': cb_purchases,
        'bbg_extras': bbg_extras,
    }


# ============================================================
# STEP 4: BUILD EXCEL SHEETS
# ============================================================
def build_monthly(daily, research):
    """Aggregate daily to monthly and merge with monthly research data."""
    monthly = daily.resample('MS').agg({
        col: ('sum' if col == 'GLD_Volume' else 'mean')
        for col in daily.columns if col != 'Date'
    }).round(2)
    monthly.index.name = 'Date'

    for src_key, col in [
        ('sge_premium', 'SGE_Premium_USD_oz'),
        ('dubai_premium', 'Dubai_Premium_USD_oz'),
        ('gold_etf', 'Global_Gold_ETF_Holdings_Tonnes'),
        ('india_gold_imports', 'India_Gold_Imports_USD_Bn'),
    ]:
        df = research[src_key]
        if len(df) > 0 and col in df.columns:
            s = df.set_index('Date')[[col]]
            monthly = monthly.join(s, how='left')

    if len(research['india_duty']) > 0 and 'India_Gold_Total_Duty_Pct' in research['india_duty'].columns:
        duty_ts = research['india_duty'].set_index('Date')[['India_Gold_Total_Duty_Pct']]
        duty_monthly = duty_ts.reindex(monthly.index, method='ffill')
        monthly = monthly.join(duty_monthly, how='left')

    # Join any Bloomberg-only extras
    for col_name, df_extra in research.get('bbg_extras', {}).items():
        s = df_extra.set_index('Date')[[col_name]]
        monthly = monthly.join(s, how='left')

    return monthly


def build_trade_by_partner(uae_imports, uae_exports):
    """Pivot ANNUAL trade data into partner-level columns."""
    if len(uae_imports) == 0 and len(uae_exports) == 0:
        return pd.DataFrame()

    dfs = []
    for df, prefix, val_col in [
        (uae_imports, 'Imp', 'Import_Value_USD'),
        (uae_exports, 'Exp', 'Export_Value_USD'),
    ]:
        if len(df) == 0:
            continue
        partners = df[~df['Partner'].isin(['World', 'Areas, nes'])]
        if len(partners) == 0:
            continue
        pivot = partners.pivot_table(
            index='Year', columns='Partner', values=val_col, aggfunc='sum'
        ).fillna(0)
        top15 = pivot.sum().nlargest(15).index
        pivot = pivot[top15]
        pivot.columns = [f"{prefix}_{c}" for c in pivot.columns]
        dfs.append(pivot)

    if not dfs:
        return pd.DataFrame()
    result = dfs[0]
    for d in dfs[1:]:
        result = result.join(d, how='outer')
    return result.fillna(0)


def build_monthly_trade(mon_imports, mon_exports):
    """Build monthly trade sheets from Comtrade monthly data."""
    if len(mon_imports) == 0 and len(mon_exports) == 0:
        return pd.DataFrame(), pd.DataFrame()

    agg_parts = []
    for df, val_col, wt_col, prefix in [
        (mon_imports, 'Import_Value_USD', 'Import_NetWeight_Kg', 'Import'),
        (mon_exports, 'Export_Value_USD', 'Export_NetWeight_Kg', 'Export'),
    ]:
        if len(df) == 0:
            continue
        rows = []
        for date, grp in df.groupby('Date'):
            world_row = grp[grp['Partner'] == 'World']
            if len(world_row) > 0:
                rows.append({
                    'Date': date,
                    f'Monthly_{prefix}_Value_USD': world_row[val_col].sum(),
                    f'Monthly_{prefix}_Weight_Kg': world_row[wt_col].sum(),
                })
            else:
                partners = grp[~grp['Partner'].isin(['World', 'Areas, nes'])]
                rows.append({
                    'Date': date,
                    f'Monthly_{prefix}_Value_USD': partners[val_col].sum(),
                    f'Monthly_{prefix}_Weight_Kg': partners[wt_col].sum(),
                })
        agg = pd.DataFrame(rows).set_index('Date')
        agg[f'Monthly_{prefix}_Weight_Tonnes'] = agg[f'Monthly_{prefix}_Weight_Kg'] / 1000
        agg_parts.append(agg)

    if agg_parts:
        monthly_agg = agg_parts[0]
        for p in agg_parts[1:]:
            monthly_agg = monthly_agg.join(p, how='outer')
        monthly_agg = monthly_agg.sort_index()
        if 'Monthly_Import_Value_USD' in monthly_agg.columns and 'Monthly_Export_Value_USD' in monthly_agg.columns:
            monthly_agg['Monthly_Net_Trade_Value_USD'] = (
                monthly_agg['Monthly_Import_Value_USD'] - monthly_agg['Monthly_Export_Value_USD']
            )
    else:
        monthly_agg = pd.DataFrame()

    partner_dfs = []
    for df, prefix, val_col in [
        (mon_imports, 'Imp', 'Import_Value_USD'),
        (mon_exports, 'Exp', 'Export_Value_USD'),
    ]:
        if len(df) == 0:
            continue
        partners = df[~df['Partner'].isin(['World', 'Areas, nes'])]
        if len(partners) == 0:
            continue
        totals = partners.groupby('Partner')[val_col].sum().nlargest(15)
        top15 = totals.index.tolist()
        filtered = partners[partners['Partner'].isin(top15)]
        pivot = filtered.pivot_table(
            index='Date', columns='Partner', values=val_col, aggfunc='sum'
        ).fillna(0)
        pivot.columns = [f"{prefix}_{c}" for c in pivot.columns]
        partner_dfs.append(pivot)

    if partner_dfs:
        monthly_partner = partner_dfs[0]
        for p in partner_dfs[1:]:
            monthly_partner = monthly_partner.join(p, how='outer')
        monthly_partner = monthly_partner.fillna(0).sort_index()
    else:
        monthly_partner = pd.DataFrame()

    return monthly_agg, monthly_partner


def build_annual_aggregate(uae_imports, uae_exports, research):
    """Build annual aggregate trade + macro data."""
    agg_dfs = []
    for df, prefix in [(uae_imports, 'Import'), (uae_exports, 'Export')]:
        if len(df) == 0:
            continue
        val_col = f'{prefix}_Value_USD'
        wt_col = f'{prefix}_NetWeight_Kg'

        yearly_rows = []
        for year, grp in df.groupby('Year'):
            world_row = grp[grp['Partner'] == 'World']
            if len(world_row) > 0:
                yearly_rows.append({
                    'Year': year,
                    val_col: world_row[val_col].sum(),
                    wt_col: world_row[wt_col].sum(),
                })
            else:
                partners = grp[~grp['Partner'].isin(['World', 'Areas, nes'])]
                yearly_rows.append({
                    'Year': year,
                    val_col: partners[val_col].sum(),
                    wt_col: partners[wt_col].sum(),
                })
        yearly = pd.DataFrame(yearly_rows).set_index('Year')
        yearly = yearly.rename(columns={
            val_col: f'Total_{prefix}_Value_USD',
            wt_col: f'Total_{prefix}_Weight_Kg',
        })
        agg_dfs.append(yearly)

    if not agg_dfs:
        yearly_trade = pd.DataFrame(index=pd.Index(range(2016, 2026), name='Year'))
    else:
        yearly_trade = agg_dfs[0]
        for d in agg_dfs[1:]:
            yearly_trade = yearly_trade.join(d, how='outer')
        yearly_trade = yearly_trade.fillna(0)

    if 'Total_Import_Value_USD' in yearly_trade.columns and 'Total_Export_Value_USD' in yearly_trade.columns:
        yearly_trade['Net_Trade_Value_USD'] = yearly_trade['Total_Import_Value_USD'] - yearly_trade['Total_Export_Value_USD']
    if 'Total_Import_Weight_Kg' in yearly_trade.columns:
        yearly_trade['Import_Tonnes'] = yearly_trade['Total_Import_Weight_Kg'] / 1000
    if 'Total_Export_Weight_Kg' in yearly_trade.columns:
        yearly_trade['Export_Tonnes'] = yearly_trade['Total_Export_Weight_Kg'] / 1000

    # Join annual research data
    for key in ['swiss_to_uae', 'turkey_uae', 'cb_purchases']:
        src = research[key]
        if len(src) > 0 and 'Year' in src.columns:
            src = src.set_index('Year').drop(columns=['Source'], errors='ignore')
            yearly_trade = yearly_trade.join(src, how='outer')

    # UAE reserves by year
    res = research['uae_reserves']
    if len(res) > 0 and 'Date' in res.columns:
        res = res.copy()
        res['Year'] = pd.to_datetime(res['Date']).dt.year
        res = res.groupby('Year').last()[['UAE_CB_Gold_Reserves_Tonnes', 'UAE_CB_Gold_Reserves_USD_Bn']]
        yearly_trade = yearly_trade.join(res, how='outer')

    # Africa pivot
    africa = research['africa_uae']
    if len(africa) > 0 and 'Year' in africa.columns:
        africa_pivot = africa.pivot_table(index='Year', columns='Country', values='Gold_Export_to_UAE_USD_Bn', aggfunc='sum')
        africa_pivot.columns = [f"Africa_{c}_to_UAE_USD_Bn" for c in africa_pivot.columns]
        yearly_trade = yearly_trade.join(africa_pivot, how='outer')

    return yearly_trade


def build_data_dictionary(use_bloomberg=False):
    """Create the data dictionary DataFrame."""
    rows = [
        ('COMEX_Gold_Close_USD', 'COMEX Gold Futures closing price (USD/oz)', 'Daily',
         'Bloomberg (GC1 Comdty)' if use_bloomberg else 'Yahoo Finance (GC=F)'),
        ('Silver_Close_USD', 'COMEX Silver Futures closing price (USD/oz)', 'Daily',
         'Bloomberg (SI1 Comdty)' if use_bloomberg else 'Yahoo Finance (SI=F)'),
        ('USD_INR', 'US Dollar to Indian Rupee exchange rate', 'Daily',
         'Bloomberg (USDINR Curncy)' if use_bloomberg else 'Yahoo Finance (INR=X)'),
        ('USD_CNY', 'US Dollar to Chinese Yuan exchange rate', 'Daily',
         'Bloomberg (USDCNY Curncy)' if use_bloomberg else 'Yahoo Finance (CNY=X)'),
        ('USD_TRY', 'US Dollar to Turkish Lira exchange rate', 'Daily',
         'Bloomberg (USDTRY Curncy)' if use_bloomberg else 'Yahoo Finance (TRY=X)'),
        ('DXY_Index', 'US Dollar Index - trade-weighted USD value', 'Daily',
         'Bloomberg (DXY Curncy)' if use_bloomberg else 'Yahoo Finance (DX-Y.NYB)'),
        ('GLD_Close', 'SPDR Gold Shares ETF closing price', 'Daily',
         'Bloomberg (GLD US Equity)' if use_bloomberg else 'Yahoo Finance (GLD)'),
        ('GLD_Volume', 'GLD daily trading volume (shares)', 'Daily',
         'Bloomberg (GLD US Equity)' if use_bloomberg else 'Yahoo Finance (GLD)'),
        ('US_10Y_Yield', 'US 10-Year Treasury Yield (%)', 'Daily',
         'Bloomberg (USGG10YR Index)' if use_bloomberg else 'Yahoo Finance (^TNX)'),
        ('VIX', 'CBOE Volatility Index', 'Daily',
         'Bloomberg (VIX Index)' if use_bloomberg else 'Yahoo Finance (^VIX)'),
        ('WTI_Crude_USD', 'WTI Crude Oil Futures closing price (USD/bbl)', 'Daily',
         'Bloomberg (CL1 Comdty)' if use_bloomberg else 'Yahoo Finance (CL=F)'),
        ('Gold_Silver_Ratio', 'Gold/Silver price ratio', 'Daily', 'Computed'),
        ('Gold_AED_oz', 'Gold price in AED/oz',
         'Daily', 'Bloomberg (XAUAED Curncy)' if use_bloomberg else 'Computed (3.6725 peg)'),
        ('Gold_INR_per_10g', 'Gold price in INR per 10 grams', 'Daily', 'Computed'),
        ('SGE_Premium_USD_oz', 'Shanghai Gold Exchange premium over London spot (USD/oz)',
         'Monthly', 'Bloomberg (GLDPSGP Index)' if use_bloomberg else 'data/sge_premium_estimate.csv'),
        ('Dubai_Premium_USD_oz', 'Dubai gold premium/discount vs London spot (USD/oz)',
         'Monthly', 'Bloomberg (GLDPDXB Index)' if use_bloomberg else 'data/dubai_premium_estimate.csv'),
        ('Global_Gold_ETF_Holdings_Tonnes', 'Global Gold ETF total holdings (tonnes)',
         'Monthly', 'Bloomberg (TGOLDTOT Index)' if use_bloomberg else 'data/gold_etf_holdings_estimate.csv'),
        ('India_Gold_Imports_USD_Bn', 'India monthly gold import value (USD billions)',
         'Monthly', 'Bloomberg (INGDIMPM Index)' if use_bloomberg else 'data/india_gold_imports_estimate.csv'),
        ('India_Gold_Total_Duty_Pct', 'India total gold import duty rate (%)', 'Event-based', 'data/india_gold_duty.csv'),
        ('UAE_CB_Gold_Reserves_Tonnes', 'UAE Central Bank gold reserves (tonnes)', 'Semi-annual', 'data/uae_cb_gold_reserves.csv'),
        ('Global_CB_Gold_Purchases_Tonnes', 'Global central bank net gold purchases (annual, tonnes)', 'Annual', 'data/global_cb_gold_purchases.csv'),
        ('Swiss_Gold_Export_to_UAE_Tonnes', 'Swiss gold exports to UAE (annual, tonnes)', 'Annual', 'data/swiss_gold_to_uae.csv'),
        ('Turkey_Gold_Export_to_UAE_USD_Bn', 'Turkey gold exports to UAE (annual, USD billions)', 'Annual', 'data/turkey_uae_gold.csv'),
        ('Total_Import_Value_USD', 'UAE total gold imports value (annual, USD) - HS 7108', 'Annual', 'UN Comtrade HS 7108'),
        ('Total_Import_Weight_Kg', 'UAE total gold imports weight (annual, Kg) - HS 7108', 'Annual', 'UN Comtrade HS 7108'),
        ('Total_Export_Value_USD', 'UAE total gold exports value (annual, USD) - HS 7108', 'Annual', 'UN Comtrade HS 7108'),
        ('Total_Export_Weight_Kg', 'UAE total gold exports weight (annual, Kg) - HS 7108', 'Annual', 'UN Comtrade HS 7108'),
    ]

    if use_bloomberg:
        rows.extend([
            ('LBMA_Gold_AM_USD', 'LBMA Gold AM Fix (USD/oz)', 'Daily', 'Bloomberg (GOLDLNAM Index)'),
            ('LBMA_Gold_PM_USD', 'LBMA Gold PM Fix (USD/oz)', 'Daily', 'Bloomberg (GOLDLNPM Index)'),
            ('DGCX_Gold_Close_USD', 'DGCX Gold Futures closing price (USD/oz)', 'Daily', 'Bloomberg (OG1 DGCX Comdty)'),
            ('Brent_Crude_USD', 'Brent Crude Oil Futures closing price (USD/bbl)', 'Daily', 'Bloomberg (CO1 Comdty)'),
            ('USD_AED', 'USD/AED exchange rate', 'Daily', 'Bloomberg (USDAED Curncy)'),
            ('US_2Y_Yield', 'US 2-Year Treasury Yield (%)', 'Daily', 'Bloomberg (USGG2YR Index)'),
            ('US_Real_Rate_10Y', 'US 10-Year Real Rate / TIPS Yield (%)', 'Daily', 'Bloomberg (USGGT10Y Index)'),
            ('Fed_Funds_Rate', 'Federal Funds Effective Rate (%)', 'Daily', 'Bloomberg (FDTRMID Index)'),
            ('US_Yield_Curve_2s10s', '2s10s yield curve spread', 'Daily', 'Computed (10Y - 2Y)'),
            ('LBMA_COMEX_Spread', 'LBMA PM Fix minus COMEX close', 'Daily', 'Computed'),
            ('Global_Gold_Mine_Supply_Tonnes', 'Global gold mine supply (tonnes)', 'Monthly', 'Bloomberg (GLDMNSUP Index)'),
            ('Gold_COMEX_Open_Interest', 'COMEX gold open interest (contracts)', 'Monthly', 'Bloomberg (GCOINTNR Index)'),
            ('Gold_CFTC_Net_Long', 'CFTC net speculative long gold (contracts)', 'Monthly', 'Bloomberg (CFAUNETL Index)'),
        ])

    variables, descs, freqs, sources = zip(*rows)
    return pd.DataFrame({
        'Variable': variables, 'Description': descs, 'Frequency': freqs, 'Source': sources,
    })


# ============================================================
# STEP 5: WRITE & FORMAT EXCEL
# ============================================================
def write_excel(output_path, daily, monthly, trade_by_partner, yearly_agg,
                india_duty, data_dict, monthly_trade_agg=None, monthly_trade_partner=None,
                daily_source='Yahoo Finance'):
    """Write all sheets and apply professional formatting."""
    print("\n" + "=" * 60)
    print("STEP 5: WRITING EXCEL WORKBOOK")
    print("=" * 60)

    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        daily.to_excel(writer, sheet_name='Daily_Market_Data')
        monthly.to_excel(writer, sheet_name='Monthly_Data')
        if monthly_trade_agg is not None and len(monthly_trade_agg) > 0:
            monthly_trade_agg.to_excel(writer, sheet_name='Monthly_Trade_Totals')
        if monthly_trade_partner is not None and len(monthly_trade_partner) > 0:
            monthly_trade_partner.to_excel(writer, sheet_name='Monthly_Trade_Partners')
        if len(trade_by_partner) > 0:
            trade_by_partner.to_excel(writer, sheet_name='Annual_Trade_By_Partner')
        yearly_agg.to_excel(writer, sheet_name='UAE_Annual_Aggregate')
        india_duty.to_excel(writer, sheet_name='India_Duty_Timeline', index=False)
        data_dict.to_excel(writer, sheet_name='Data_Dictionary', index=False)

    print("  Data written. Formatting...")

    wb = load_workbook(output_path)
    hdr_font = Font(name='Arial', bold=True, size=10, color='FFFFFF')
    hdr_fill = PatternFill('solid', fgColor='1F4E79')
    data_font = Font(name='Arial', size=9)
    border = Border(bottom=Side(style='thin', color='D9D9D9'))

    for ws_name in wb.sheetnames:
        ws = wb[ws_name]
        for cell in ws[1]:
            cell.font = hdr_font
            cell.fill = hdr_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
        for col in ws.columns:
            col_letter = get_column_letter(col[0].column)
            max_len = max((len(str(c.value or '')) for c in col), default=10)
            ws.column_dimensions[col_letter].width = min(max_len + 3, 28)
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.font = data_font
                cell.border = border
        ws.freeze_panes = 'B2'

    # Cover sheet
    ws_cover = wb.create_sheet('Summary', 0)
    mt_agg_len = len(monthly_trade_agg) if monthly_trade_agg is not None else 0
    mt_part_len = len(monthly_trade_partner) if monthly_trade_partner is not None else 0
    cover_rows = [
        ['UAE GOLD TRADE & DUBAI PREMIUM - HISTORICAL DATABASE'],
        [''],
        ['Created:', datetime.now().strftime('%Y-%m-%d')],
        ['Market Data Source:', daily_source],
        ['Coverage:', f'{daily.index.min().strftime("%Y-%m-%d")} to {daily.index.max().strftime("%Y-%m-%d")}'],
        [''],
        ['SHEET GUIDE:'],
        ['Sheet', 'Description', 'Frequency', 'Records'],
        ['Daily_Market_Data', 'Gold, FX, VIX, DXY, oil, yields, GLD ETF', 'Daily', len(daily)],
        ['Monthly_Data', 'Monthly avgs + Dubai/SGE premium, ETF, India imports, duty', 'Monthly', len(monthly)],
        ['Monthly_Trade_Totals', 'UAE gold import/export totals from Comtrade (HS 7108)', 'Monthly', mt_agg_len],
        ['Monthly_Trade_Partners', 'UAE gold trade by top 15 partners from Comtrade (HS 7108)', 'Monthly', mt_part_len],
        ['Annual_Trade_By_Partner', 'UAE gold imports/exports by top 15 partners (HS 7108)', 'Annual', len(trade_by_partner)],
        ['UAE_Annual_Aggregate', 'Total trade + Swiss/Turkey/Africa + CB reserves', 'Annual', len(yearly_agg)],
        ['India_Duty_Timeline', 'India gold import duty changes (BCD + AIDC)', 'Event', len(india_duty)],
        ['Data_Dictionary', 'Variable descriptions, sources, frequencies', '-', len(data_dict)],
        [''],
        ['KEY DRIVERS OF DUBAI GOLD PREMIUM/DISCOUNT:'],
        ['1. Dubai Premium vs London (monthly)'],
        ['2. Shanghai Gold Exchange Premium (monthly)'],
        ['3. India Gold Import Duty (event-based)'],
        ['4. FX: USD/INR, USD/CNY, USD/TRY (daily)'],
        ['5. DXY US Dollar Index (daily)'],
        ['6. US 10Y Treasury Yield (daily)'],
        ['7. VIX - risk sentiment (daily)'],
        ['8. Global Gold ETF Holdings (monthly)'],
        ['9. India Gold Imports (monthly)'],
        ['10. UAE CB Gold Reserves (semi-annual)'],
        ['11. Global CB Gold Purchases (annual)'],
        ['12. Swiss Gold Exports to UAE (annual)'],
        ['13. Turkey-UAE Gold Trade (annual)'],
        ['14. African Gold to UAE by country (annual)'],
        ['15. Gold/Silver Ratio (daily)'],
        ['16. WTI Crude Oil (daily)'],
        ['17. GLD ETF Price & Volume (daily)'],
        [''],
        ['DATA QUALITY NOTES:'],
        [f'- Daily market data: {daily_source}'],
        ['- UN Comtrade trade: Official government-reported statistics'],
        ['- Research data loaded from CSV files in data/ folder (editable)'],
        ['- Bloomberg mode replaces estimates with terminal data where available'],
        ['- Annual flows (Swiss, Turkey, Africa): Estimates from multiple public sources'],
        ['- India duty: Official Union Budget rates'],
        ['- Bloomberg mode adds: LBMA fixes, DGCX, real rates, CFTC positioning, mine supply'],
    ]
    for r in cover_rows:
        ws_cover.append(r)

    ws_cover['A1'].font = Font(name='Arial', bold=True, size=16, color='1F4E79')
    ws_cover.merge_cells('A1:D1')
    for cell in ws_cover[8]:
        cell.font = hdr_font
        cell.fill = hdr_fill
    ws_cover.column_dimensions['A'].width = 28
    ws_cover.column_dimensions['B'].width = 60
    ws_cover.column_dimensions['C'].width = 15
    ws_cover.column_dimensions['D'].width = 10

    wb.save(output_path)
    size_mb = os.path.getsize(output_path) / 1024 / 1024
    print(f"  => Saved: {output_path} ({size_mb:.1f} MB)")


# ============================================================
# MAIN
# ============================================================
def main():
    parser = argparse.ArgumentParser(description='UAE Gold Trade & Dubai Premium Historical Data Collector')
    parser.add_argument('--start', default=DEFAULT_START, help=f'Start date (default: {DEFAULT_START})')
    parser.add_argument('--end', default=DEFAULT_END, help=f'End date (default: today)')
    parser.add_argument('--output', '-o', default=DEFAULT_OUTPUT, help=f'Output Excel path (default: {DEFAULT_OUTPUT})')
    parser.add_argument('--skip-comtrade', action='store_true', help='Skip UN Comtrade API calls (use if rate-limited)')
    parser.add_argument('--bloomberg', action='store_true',
                        help='Use Bloomberg Terminal (blpapi) for market data instead of Yahoo Finance. '
                             'Requires Bloomberg Terminal running with DAPI enabled and blpapi installed.')
    parser.add_argument('--proxy', default=None,
                        help='Proxy URL for all HTTP requests (yfinance, Comtrade, etc.). '
                             'Examples: http://127.0.0.1:8080, socks5://127.0.0.1:1080, '
                             'http://user:pass@proxy.corp.com:3128')
    parser.add_argument('--data-dir', default=DEFAULT_DATA_DIR,
                        help=f'Path to directory containing research CSV files (default: ./data)')
    args = parser.parse_args()

    start_year = int(args.start[:4])
    end_year = int(args.end[:4])

    print(f"\nUAE Gold Trade Historical Data Collector")
    print(f"Period: {args.start} to {args.end}")
    print(f"Output: {args.output}")
    print(f"Mode:   {'Bloomberg Terminal' if args.bloomberg else 'Yahoo Finance (default)'}")
    print(f"Data:   {args.data_dir}")
    if args.proxy:
        print(f"Proxy:  {args.proxy}")
    print(f"{'=' * 60}")

    # Configure proxy
    configure_proxy(args.proxy)

    # Validate data directory
    if not os.path.isdir(args.data_dir):
        print(f"\n  WARNING: Data directory not found: {args.data_dir}")
        print(f"  Research data will be empty. Create the directory with CSV files.")
        print(f"  Expected files: {', '.join(DATA_FILES.keys())}")
    else:
        missing = [f for f in DATA_FILES if not os.path.exists(os.path.join(args.data_dir, f))]
        if missing:
            print(f"\n  WARNING: Missing data files: {', '.join(missing)}")

    # Step 1: Daily market data
    daily_source = 'Yahoo Finance'
    bbg_monthly_data = {}

    if args.bloomberg:
        try:
            _ensure_blpapi()
            daily, daily_source = collect_daily_market_data_bloomberg(args.start, args.end)
            bbg_monthly_data = collect_monthly_bloomberg_data(args.start, args.end)
        except ImportError:
            print("\n  ERROR: blpapi not installed. Install with: pip install blpapi")
            print("  Falling back to Yahoo Finance.")
            daily = collect_daily_market_data(args.start, args.end)
        except Exception as e:
            print(f"\n  Bloomberg error: {e}")
            print("  Falling back to Yahoo Finance.")
            daily = collect_daily_market_data(args.start, args.end)
    else:
        daily = collect_daily_market_data(args.start, args.end)

    # Step 2: UN Comtrade trade data
    if args.skip_comtrade:
        print("\n  Skipping Comtrade (--skip-comtrade flag)")
        ann_imports, ann_exports = pd.DataFrame(), pd.DataFrame()
        mon_imports, mon_exports = pd.DataFrame(), pd.DataFrame()
    else:
        ann_imports, ann_exports, mon_imports, mon_exports = collect_comtrade_data(start_year, end_year)

    # Step 3: Research-based data
    research = collect_research_data(args.data_dir, bbg_monthly=bbg_monthly_data)

    # Step 4: Build sheets
    print("\n" + "=" * 60)
    print("STEP 4: BUILDING EXCEL SHEETS")
    print("=" * 60)

    monthly = build_monthly(daily, research)
    print(f"  Monthly market data: {monthly.shape}")

    monthly_trade_agg, monthly_trade_partner = build_monthly_trade(mon_imports, mon_exports)
    print(f"  Monthly trade totals: {monthly_trade_agg.shape if len(monthly_trade_agg) > 0 else '(empty)'}")
    print(f"  Monthly trade by partner: {monthly_trade_partner.shape if len(monthly_trade_partner) > 0 else '(empty)'}")

    trade_by_partner = build_trade_by_partner(ann_imports, ann_exports)
    print(f"  Annual trade by partner: {trade_by_partner.shape}")

    yearly_agg = build_annual_aggregate(ann_imports, ann_exports, research)
    print(f"  Annual aggregate: {yearly_agg.shape}")

    data_dict = build_data_dictionary(use_bloomberg=args.bloomberg)

    # Step 5: Write Excel
    write_excel(
        args.output, daily, monthly, trade_by_partner, yearly_agg,
        research['india_duty'], data_dict,
        monthly_trade_agg=monthly_trade_agg,
        monthly_trade_partner=monthly_trade_partner,
        daily_source=daily_source,
    )

    print(f"\n{'=' * 60}")
    print("DONE! All data collected and saved.")
    if args.bloomberg:
        print(f"Data source: {daily_source}")
        if bbg_monthly_data:
            print(f"Bloomberg monthly series sourced: {list(bbg_monthly_data.keys())}")
    print(f"{'=' * 60}")


if __name__ == '__main__':
    main()
